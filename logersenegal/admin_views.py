import openpyxl
from openpyxl.styles import Font, Alignment
from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from django.db.models import Count, Q
from users.models import User, KYCProfile
from solvable.models import RentalFiliation, IncidentReport, PaymentHistory
import datetime
from django.db.models import Sum
from logersn.models import Property, Transaction

@staff_member_required
def admin_statistics_view(request):
    # 1. User Stats
    total_users = User.objects.count()
    users_by_role = User.objects.values('role').annotate(count=Count('id'))
    verified_pros = User.objects.filter(is_verified_pro=True).count()
    kyc_pending = KYCProfile.objects.filter(vision_api_status='PENDING').count()

    # 2. Property Stats
    total_properties = Property.objects.count()
    published_properties = Property.objects.filter(is_published=True).count()
    
    # 3. Rental Stats
    total_filiations = RentalFiliation.objects.count()
    active_filiations = RentalFiliation.objects.filter(status='ACTIVE').count()
    
    # NEW KPI: Occupancy Rate
    properties_occupied = RentalFiliation.objects.filter(status='ACTIVE').values('property').distinct().count()
    occupancy_rate = (properties_occupied / total_properties * 100) if total_properties > 0 else 0

    # 4. Incident & Contestation Stats
    total_incidents = IncidentReport.objects.count()
    contested_incidents = IncidentReport.objects.filter(is_contested=True).count()
    resolved_incidents = IncidentReport.objects.filter(status='RESOLVED').count()
    
    contestation_rate = (contested_incidents / total_incidents * 100) if total_incidents > 0 else 0
    resolution_rate = (resolved_incidents / total_incidents * 100) if total_incidents > 0 else 0

    # 5. Payment Stats & Recovery
    total_payments = PaymentHistory.objects.count()
    paid_payments = PaymentHistory.objects.filter(status='PAID').count()
    late_payments = PaymentHistory.objects.filter(status='LATE').count()
    unpaid_payments = PaymentHistory.objects.filter(status='UNPAID').count()
    
    # NEW KPI: Collection/Recovery Rate
    collection_rate = (paid_payments / total_payments * 100) if total_payments > 0 else 0
    
    # NEW KPI: KYC & Pro Verification
    verified_kyc_count = KYCProfile.objects.filter(vision_api_status='APPROVED').count()
    kyc_compliance_rate = (verified_kyc_count / total_users * 100) if total_users > 0 else 0
    
    total_pros = User.objects.exclude(role='TENANT').count()
    pro_verification_rate = (verified_pros / total_pros * 100) if total_pros > 0 else 0
    
    # 7. FINANCIAL STATS (DigitalH PÉAGE)
    total_revenue = Transaction.objects.filter(status='SUCCESS').aggregate(total=Sum('amount'))['total'] or 0
    revenue_publication = Transaction.objects.filter(status='SUCCESS', transaction_type='PUBLICATION').aggregate(total=Sum('amount'))['total'] or 0
    revenue_boost = Transaction.objects.filter(status='SUCCESS', transaction_type='BOOST').aggregate(total=Sum('amount'))['total'] or 0
    revenue_popup = Transaction.objects.filter(status='SUCCESS', transaction_type='POPUP').aggregate(total=Sum('amount'))['total'] or 0
    
    # Revenue this month
    this_month_start = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    month_revenue = Transaction.objects.filter(status='SUCCESS', created_at__gte=this_month_start).aggregate(total=Sum('amount'))['total'] or 0
    
    # 6. Time Series (Last 6 months)
    today = timezone.now()
    months = []
    user_growth = []
    contract_growth = []
    for i in range(5, -1, -1):
        month_start = (today - datetime.timedelta(days=i*30)).replace(day=1)
        months.append(month_start.strftime('%B'))
        user_growth.append(User.objects.filter(date_joined__lte=month_start + datetime.timedelta(days=30)).count())
        contract_growth.append(RentalFiliation.objects.filter(created_at__lte=month_start + datetime.timedelta(days=30)).count())

    # Handle Excel Export
    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Statistiques Solvable"
        
        # Headers
        ws.append(["Indicateur", "Valeur"])
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        
        ws.append(["Utilisateurs Totaux", total_users])
        ws.append(["Professionnels Vérifiés", verified_pros])
        ws.append(["Taux de Vérification Pro (%)", f"{pro_verification_rate:.2f}"])
        ws.append(["Annonces Publiées", published_properties])
        ws.append(["Taux d'Occupation (%)", f"{occupancy_rate:.2f}"])
        ws.append(["Contrats Actifs", active_filiations])
        ws.append(["Taux de Contestation (%)", f"{contestation_rate:.2f}"])
        ws.append(["Taux de Résolution Litiges (%)", f"{resolution_rate:.2f}"])
        ws.append(["Paiements à temps", paid_payments])
        ws.append(["Taux de Recouvrement (%)", f"{collection_rate:.2f}"])
        ws.append(["Conformité KYC (%)", f"{kyc_compliance_rate:.2f}"])
        ws.append(["Paiements Impayés", unpaid_payments])
        
        # Breakdown by role
        ws.append([])
        ws.append(["Rôle", "Nombre d'inscrits"])
        for stat in users_by_role:
            ws.append([stat['role'], stat['count']])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=stats_solvable_{today.strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response

    context = {
        'total_users': total_users,
        'users_by_role': list(users_by_role),
        'verified_pros': verified_pros,
        'kyc_pending': kyc_pending,
        'total_properties': total_properties,
        'published_properties': published_properties,
        'total_filiations': total_filiations,
        'active_filiations': active_filiations,
        'total_incidents': total_incidents,
        'contestation_rate': round(contestation_rate, 1),
        'resolution_rate': round(resolution_rate, 1),
        'collection_rate': round(collection_rate, 1),
        'occupancy_rate': round(occupancy_rate, 1),
        'kyc_compliance_rate': round(kyc_compliance_rate, 1),
        'pro_verification_rate': round(pro_verification_rate, 1),
        'total_payments': total_payments,
        'paid_payments': paid_payments,
        'late_payments': late_payments,
        'unpaid_payments': unpaid_payments,
        'labels_months': months,
        'total_revenue': total_revenue,
        'revenue_publication': revenue_publication,
        'revenue_boost': revenue_boost,
        'revenue_popup': revenue_popup,
        'month_revenue': month_revenue,
        'data_users': user_growth,
        'data_contracts': contract_growth,
    }
    
    return render(request, 'admin/statistics.html', context)
