import openpyxl
from openpyxl.styles import Font, Alignment
from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from django.db.models import Count, Q
from users.models import User, KYCProfile

import datetime
from django.db.models import Sum
from django.utils import timezone
from logersn.models import Property, Transaction

@staff_member_required
def admin_statistics_view(request):
    today = timezone.now()
    try:
        # 1. User Stats
        total_users = User.objects.count()
        users_by_role = User.objects.values('role').annotate(count=Count('id'))
        verified_pros = User.objects.filter(is_verified_pro=True).count()
        kyc_pending = KYCProfile.objects.filter(vision_api_status='PENDING').count()

        # 2. Property Stats
        total_properties = Property.objects.count()
        published_properties = Property.objects.filter(is_published=True).count()
        
        # 3. Rental Stats (Removed)
        total_filiations = 0
        active_filiations = 0
        properties_occupied = 0
        occupancy_rate = 0

        # 4. Incident & Contestation Stats (Removed)
        total_incidents = 0
        contested_incidents = 0
        resolved_incidents = 0
        contestation_rate = 0
        resolution_rate = 0

        # 5. Payment Stats & Recovery (Removed)
        total_payments = 0
        paid_payments = 0
        late_payments = 0
        unpaid_payments = 0
        collection_rate = 0
        
        # NEW KPI: KYC & Pro Verification
        verified_kyc_count = KYCProfile.objects.filter(vision_api_status='APPROVED').count()
        kyc_compliance_rate = (verified_kyc_count / total_users * 100) if total_users > 0 else 0
        
        total_pros = User.objects.exclude(role='TENANT').count()
        pro_verification_rate = (verified_pros / total_pros * 100) if total_pros > 0 else 0
        
        # 7. FINANCIAL STATS (DigitalH PÉAGE)
        total_revenue = Transaction.objects.filter(status='SUCCESS').aggregate(total=Sum('amount'))['total'] or 0
        revenue_publication = Transaction.objects.filter(status='SUCCESS', transaction_type='PUBLICATION').aggregate(total=Sum('amount'))['total'] or 0
        revenue_boost = Transaction.objects.filter(status='SUCCESS', transaction_type='BOOST').aggregate(total=Sum('amount'))['total'] or 0
        revenue_popup = Transaction.objects.filter(status='SUCCESS', transaction_type='POPUP').aggregate(total=Sum('amount'))['total'] or 0
        
        # Revenue this month
        this_month_start = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        month_revenue = Transaction.objects.filter(status='SUCCESS', created_at__gte=this_month_start).aggregate(total=Sum('amount'))['total'] or 0
        
        # 6. Time Series (Last 6 months)
        months = []
        user_growth = []
        contract_growth = []
        for i in range(5, -1, -1):
            month_start = (today - datetime.timedelta(days=i*30)).replace(day=1)
            months.append(month_start.strftime('%B'))
            user_growth.append(User.objects.filter(date_joined__lte=month_start + datetime.timedelta(days=30)).count())
            contract_growth.append(0)
    except Exception as e:
        # Fallback values for all stats to avoid 500, but with a warning for staff
        total_users = verified_pros = kyc_pending = total_properties = published_properties = 0
        total_filiations = active_filiations = occupancy_rate = 0
        total_incidents = contested_incidents = resolved_incidents = contestation_rate = resolution_rate = 0
        total_payments = paid_payments = late_payments = unpaid_payments = collection_rate = 0
        kyc_compliance_rate = total_pros = pro_verification_rate = 0
        total_revenue = revenue_publication = revenue_boost = revenue_popup = month_revenue = 0
        months = user_growth = contract_growth = []
        users_by_role = []
        import logging
        logging.error(f"Admin Statistics 500 prevention alert: {e}")

    # Handle Excel Export
    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Statistiques Solvable"
        
        # Headers
        ws.append(["Indicateur", "Valeur"])
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        
        ws.append(["Utilisateurs Totaux", total_users])
        ws.append(["Professionnels Vérifiés", verified_pros])
        ws.append(["Taux de Vérification Pro (%)", f"{pro_verification_rate:.2f}"])
        ws.append(["Annonces Publiées", published_properties])
        ws.append(["Taux d'Occupation (%)", f"{occupancy_rate:.2f}"])
        ws.append(["Contrats Actifs", active_filiations])
        ws.append(["Taux de Contestation (%)", f"{contestation_rate:.2f}"])
        ws.append(["Taux de Résolution Litiges (%)", f"{resolution_rate:.2f}"])
        ws.append(["Paiements à temps", paid_payments])
        ws.append(["Taux de Recouvrement (%)", f"{collection_rate:.2f}"])
        ws.append(["Conformité KYC (%)", f"{kyc_compliance_rate:.2f}"])
        ws.append(["Paiements Impayés", unpaid_payments])
        
        # Breakdown by role
        ws.append([])
        ws.append(["Rôle", "Nombre d'inscrits"])
        for stat in users_by_role:
            ws.append([stat['role'], stat['count']])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=stats_logertogo_{today.strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response

    context = {
        'total_users': total_users,
        'users_by_role': list(users_by_role),
        'verified_pros': verified_pros,
        'kyc_pending': kyc_pending,
        'total_properties': total_properties,
        'published_properties': published_properties,
        'total_filiations': total_filiations,
        'active_filiations': active_filiations,
        'total_incidents': total_incidents,
        'contestation_rate': round(contestation_rate, 1),
        'resolution_rate': round(resolution_rate, 1),
        'collection_rate': round(collection_rate, 1),
        'occupancy_rate': round(occupancy_rate, 1),
        'kyc_compliance_rate': round(kyc_compliance_rate, 1),
        'pro_verification_rate': round(pro_verification_rate, 1),
        'total_payments': total_payments,
        'paid_payments': paid_payments,
        'late_payments': late_payments,
        'unpaid_payments': unpaid_payments,
        'labels_months': months,
        'total_revenue': total_revenue,
        'revenue_publication': revenue_publication,
        'revenue_boost': revenue_boost,
        'revenue_popup': revenue_popup,
        'month_revenue': month_revenue,
        'data_users': user_growth,
        'data_contracts': contract_growth,
    }
    
    return render(request, 'admin/statistics.html', context)

@staff_member_required
def admin_marketing_email_view(request):
    from django.shortcuts import redirect
    from django.contrib import messages
    from django.core.mail import EmailMultiAlternatives
    from django.conf import settings
    from users.models import User

    user_ids = request.session.get('marketing_user_ids', [])
    if not user_ids:
        messages.error(request, "Aucun utilisateur sélectionné pour la campagne.")
        return redirect('admin:users_user_changelist')
    
    users = User.objects.filter(id__in=user_ids)
    
    if request.method == 'POST':
        subject = request.POST.get('subject')
        message_content = request.POST.get('message')
        is_html = request.POST.get('is_html') == 'on'
        
        count = 0
        for user in users:
            if user.email:
                # On remplace les tags [NOM], [PRENOM] si présents
                personalized_message = message_content.replace('[PRENOM]', user.first_name).replace('[NOM]', user.last_name)
                
                msg = EmailMultiAlternatives(
                    subject,
                    personalized_message if not is_html else "Contenu HTML : veuillez utiliser un client mail moderne.",
                    settings.DEFAULT_FROM_EMAIL,
                    [user.email],
                    bcc=['contact@logertg.com']
                )
                if is_html:
                    msg.attach_alternative(personalized_message, "text/html")
                
                try:
                    msg.send()
                    count += 1
                except:
                    pass
        
        messages.success(request, f"🚀 Campagne terminée : {count} e-mails envoyés avec succès.")
        if 'marketing_user_ids' in request.session:
            del request.session['marketing_user_ids']
        return redirect('admin:users_user_changelist')
        
    return render(request, 'admin/marketing_email.html', {
        'users': users,
        'count': users.count()
    })
